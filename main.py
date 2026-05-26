"""
Solar Tracker Monitor  v5.0
Real-time dashboard for Arduino-based solar tracking systems.
"""

import sys
import re
import time
import threading
import collections
import math
import random
import subprocess
from datetime import datetime

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QFrame, QPushButton, QComboBox, QSpinBox, QProgressBar,
    QFileDialog, QMessageBox, QSizePolicy, QGraphicsDropShadowEffect,
)
from PySide6.QtCore import QTimer, Signal, QObject, Qt
from PySide6.QtGui import QColor, QFont, QPalette

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.ticker as ticker

try:
    import serial
    import serial.tools.list_ports
    SERIAL_AVAILABLE = True
except ImportError:
    SERIAL_AVAILABLE = False


MAX_POINTS = 600

PATTERN = re.compile(
    r"T:\s*(\d+).*?L:\s*(\d+).*?R:\s*(\d+).*?B:\s*(\d+)"
    r".*?U=([\d.]+)V.*?I=([\d.]+)mA.*?P=([\d.]+)mW"
    r".*?rot=(\d+).*?hoek=(\d+)",
    re.IGNORECASE,
)

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "bg":       "#0D1117",
    "surface":  "#161B22",
    "border":   "#30363D",
    "blue":     "#58A6FF",
    "green":    "#3FB950",
    "orange":   "#F0883E",
    "purple":   "#BC8CFF",
    "red":      "#F85149",
    "cyan":     "#39C5CF",
    "yellow":   "#E3B341",
    "text":     "#E6EDF3",
    "muted":    "#8B949E",
    "accent":   "#1F6FEB",
    "success":  "#238636",
    "card":     "#1C2128",
}


# ─────────────────────────────────────────────────────────────────────────────
# Data model
# ─────────────────────────────────────────────────────────────────────────────
class DataModel:
    def __init__(self):
        self.timestamps  = collections.deque(maxlen=MAX_POINTS)
        self.voltages    = collections.deque(maxlen=MAX_POINTS)
        self.currents    = collections.deque(maxlen=MAX_POINTS)
        self.powers      = collections.deque(maxlen=MAX_POINTS)
        self.energies    = collections.deque(maxlen=MAX_POINTS)
        self.ldr_top     = collections.deque(maxlen=MAX_POINTS)
        self.ldr_right   = collections.deque(maxlen=MAX_POINTS)
        self.ldr_bottom  = collections.deque(maxlen=MAX_POINTS)
        self.ldr_left    = collections.deque(maxlen=MAX_POINTS)
        self.rot_pos     = collections.deque(maxlen=MAX_POINTS)
        self.hoek_pos    = collections.deque(maxlen=MAX_POINTS)

        self.all_rows     = []
        self.measure_rows = []

        self.total_energy_mWh = 0.0
        self.last_sample_time = None
        self.session_start    = None

    def add_sample(self, t_rel, top, left, right, bottom,
                   voltage, current_mA, power_mW, rot, hoek, measuring):
        now = time.time()
        if self.session_start is None:
            self.session_start = now
        if self.last_sample_time is None:
            dt = 0.0
        else:
            dt = now - self.last_sample_time
        self.last_sample_time = now

        dE = power_mW * dt / 3600.0
        self.total_energy_mWh += dE

        self.timestamps.append(t_rel)
        self.voltages.append(voltage)
        self.currents.append(current_mA)
        self.powers.append(power_mW)
        self.energies.append(self.total_energy_mWh)
        self.ldr_top.append(top)
        self.ldr_right.append(right)
        self.ldr_bottom.append(bottom)
        self.ldr_left.append(left)
        self.rot_pos.append(rot)
        self.hoek_pos.append(hoek)

        row = {
            "timestamp":         datetime.fromtimestamp(now),
            "t_rel_s":           t_rel,
            "voltage_V":         voltage,
            "current_mA":        current_mA,
            "power_mW":          power_mW,
            "energy_mWh_total":  self.total_energy_mWh,
            "ldr_top":           top,
            "ldr_right":         right,
            "ldr_bottom":        bottom,
            "ldr_left":          left,
            "rot_deg":           rot,
            "hoek_deg":          hoek,
            "dt_s":              dt,
        }
        self.all_rows.append(row)
        if measuring:
            self.measure_rows.append(row)

    def clear(self):
        for dq in [self.timestamps, self.voltages, self.currents, self.powers,
                   self.energies, self.ldr_top, self.ldr_right, self.ldr_bottom,
                   self.ldr_left, self.rot_pos, self.hoek_pos]:
            dq.clear()
        self.all_rows.clear()
        self.measure_rows.clear()
        self.total_energy_mWh = 0.0
        self.last_sample_time = None
        self.session_start    = None


class DataSignals(QObject):
    new_line = Signal(str)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────
def export_to_xlsx(rows, path):
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Measurement Data ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Measurement Data"
    ws.sheet_view.showGridLines = False

    COLS = [
        ("timestamp",        "Timestamp",        20, "YYYY-MM-DD HH:MM:SS"),
        ("t_rel_s",          "Time (s)",          10, "0.00"),
        ("voltage_V",        "Voltage (V)",       13, "0.000"),
        ("current_mA",       "Current (mA)",      14, "0.000"),
        ("power_mW",         "Power (mW)",        13, "0.000"),
        ("energy_mWh_total", "Energy (mWh)",      15, "0.00000"),
        ("ldr_top",          "LDR Top",           10, "0"),
        ("ldr_right",        "LDR Right",         11, "0"),
        ("ldr_bottom",       "LDR Bottom",        12, "0"),
        ("ldr_left",         "LDR Left",          10, "0"),
        ("rot_deg",          "Rotation (°)",      13, "0"),
        ("hoek_deg",         "Tilt (°)",          10, "0"),
    ]

    THIN   = Side(style="thin",   color="D0D7E3")
    THICK  = Side(style="medium", color="1B3A6B")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_F  = PatternFill("solid", start_color="1B3A6B")
    HDR_FT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    HDR_AL = Alignment(horizontal="center", vertical="center")
    ALT_F  = PatternFill("solid", start_color="EEF3FB")
    DAT_FT = Font(name="Calibri", size=10)
    DAT_AL = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[1].height = 22
    for c, (_, hdr, width, _fmt) in enumerate(COLS, 1):
        cell = ws.cell(1, c, hdr)
        cell.fill      = HDR_F
        cell.font      = HDR_FT
        cell.alignment = HDR_AL
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width

    n = len(rows)
    for r_i, row in enumerate(rows, 2):
        shade = ALT_F if r_i % 2 == 0 else None
        for c, (key, _, _, fmt) in enumerate(COLS, 1):
            val  = row.get(key, "")
            cell = ws.cell(r_i, c, val)
            cell.font      = DAT_FT
            cell.border    = BORDER
            cell.alignment = DAT_AL
            cell.number_format = fmt
            if shade:
                cell.fill = shade

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    # ── Sheet 2: Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    T_FONT  = Font(bold=True,  name="Calibri", size=16, color="1B3A6B")
    S_FONT  = Font(bold=False, name="Calibri", size=10, color="555555", italic=True)
    LH_FONT = Font(bold=True,  name="Calibri", size=10, color="FFFFFF")
    LD_FONT = Font(name="Calibri", size=10)
    LH_FILL = PatternFill("solid", start_color="1B3A6B")
    LA_FILL = PatternFill("solid", start_color="EEF3FB")
    BORDER2 = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws2["A1"] = "Solar Tracker — Measurement Report"
    ws2["A1"].font = T_FONT
    ws2["A2"] = f"Exported: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    ws2["A2"].font = S_FONT
    ws2.row_dimensions[3].height = 8

    stats = [
        ("Parameter",            "Value"),
        ("Total Samples",        f"=COUNTA('Measurement Data'!B2:B{n+1})"),
        ("Duration (s)",         f"=MAX('Measurement Data'!B2:B{n+1})-MIN('Measurement Data'!B2:B{n+1})"),
        ("Average Power (mW)",   f"=AVERAGE('Measurement Data'!E2:E{n+1})"),
        ("Max Power (mW)",       f"=MAX('Measurement Data'!E2:E{n+1})"),
        ("Min Power (mW)",       f"=MIN('Measurement Data'!E2:E{n+1})"),
        ("Average Voltage (V)",  f"=AVERAGE('Measurement Data'!C2:C{n+1})"),
        ("Max Voltage (V)",      f"=MAX('Measurement Data'!C2:C{n+1})"),
        ("Average Current (mA)", f"=AVERAGE('Measurement Data'!D2:D{n+1})"),
        ("Total Energy (mWh)",   f"=MAX('Measurement Data'!F2:F{n+1})"),
        ("Start Time",           rows[0]["timestamp"].strftime("%Y-%m-%d %H:%M:%S") if rows else "—"),
        ("End Time",             rows[-1]["timestamp"].strftime("%Y-%m-%d %H:%M:%S") if rows else "—"),
    ]

    for i, (label, val) in enumerate(stats):
        row_n = i + 4
        ca = ws2.cell(row_n, 1, label)
        cb = ws2.cell(row_n, 2, val)
        if i == 0:
            ca.fill = LH_FILL; cb.fill = LH_FILL
            ca.font = LH_FONT; cb.font = LH_FONT
            ca.alignment = Alignment(horizontal="center")
            cb.alignment = Alignment(horizontal="center")
        else:
            if i % 2 == 0:
                ca.fill = LA_FILL; cb.fill = LA_FILL
            ca.font = LD_FONT; cb.font = LD_FONT
            ca.border = BORDER2; cb.border = BORDER2
            cb.number_format = "0.000"
            cb.alignment = Alignment(horizontal="right")

    # ── Sheet 3: Charts ───────────────────────────────────────────────────
    wc = wb.create_sheet("Charts")
    wc.sheet_view.showGridLines = False
    # col index map (1-based in ws):
    # 1=timestamp, 2=time_s, 3=voltage, 4=current, 5=power, 6=energy
    # 7=ldr_top, 8=ldr_right, 9=ldr_bottom, 10=ldr_left, 11=rot, 12=tilt
    cats = Reference(ws, min_col=2, min_row=2, max_row=n + 1)

    def mkchart(title, y_col, y_label, color, anchor, smooth=True):
        ch = LineChart()
        ch.title  = title
        ch.style  = 10
        ch.width  = 25
        ch.height = 14
        ch.y_axis.title   = y_label
        ch.x_axis.title   = "Time (s)"
        ch.legend         = None
        ch.y_axis.numFmt  = "0.00"
        ch.x_axis.numFmt  = "0.00"
        ch.plot_area.graphicalProperties = None
        data_ref = Reference(ws, min_col=y_col, min_row=1, max_row=n + 1)
        ch.add_data(data_ref, titles_from_data=True)
        ch.set_categories(cats)
        s = ch.series[0]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width     = 22000
        s.smooth = smooth
        wc.add_chart(ch, anchor)

    mkchart("Voltage (V)",   3, "V",    "0078D4", "A1")
    mkchart("Current (mA)",  4, "mA",   "13A10E", "N1")
    mkchart("Power (mW)",    5, "mW",   "FF8C00", "A33")
    mkchart("Energy (mWh)",  6, "mWh",  "8764B8", "N33")

    # LDR 4-series
    ldr = LineChart()
    ldr.title         = "LDR Sensors"
    ldr.style         = 10
    ldr.width         = 25
    ldr.height        = 14
    ldr.y_axis.title  = "LDR Value"
    ldr.x_axis.title  = "Time (s)"
    ldr.x_axis.numFmt = "0.00"
    for col, color in zip([7, 8, 9, 10], ["D13438", "FF8C00", "13A10E", "0078D4"]):
        ref = Reference(ws, min_col=col, min_row=1, max_row=n + 1)
        ldr.add_data(ref, titles_from_data=True)
        ldr.series[-1].graphicalProperties.line.solidFill = color
        ldr.series[-1].graphicalProperties.line.width     = 16000
    ldr.set_categories(cats)
    wc.add_chart(ldr, "A65")

    # Tracker position 2-series
    pos = LineChart()
    pos.title         = "Tracker Position"
    pos.style         = 10
    pos.width         = 25
    pos.height        = 14
    pos.y_axis.title  = "Degrees (°)"
    pos.x_axis.title  = "Time (s)"
    pos.x_axis.numFmt = "0.00"
    for col, color in zip([11, 12], ["C50F1F", "00B7C3"]):
        ref = Reference(ws, min_col=col, min_row=1, max_row=n + 1)
        pos.add_data(ref, titles_from_data=True)
        pos.series[-1].graphicalProperties.line.solidFill = color
        pos.series[-1].graphicalProperties.line.width     = 16000
        pos.series[-1].smooth = True
    pos.set_categories(cats)
    wc.add_chart(pos, "N65")

    # sheet tab order
    wb.move_sheet("Summary",         offset=-2)
    wb.move_sheet("Charts",          offset=-1)

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# Metric card widget
# ─────────────────────────────────────────────────────────────────────────────
class MetricCard(QFrame):
    def __init__(self, label, unit, color=C["blue"], parent=None):
        super().__init__(parent)
        self.setObjectName("MetricCard")
        self._color = color
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(2)

        self._lbl = QLabel(label.upper())
        self._lbl.setObjectName("MetricLabel")
        self._val = QLabel("—")
        self._val.setObjectName("MetricValue")
        self._unit = QLabel(unit)
        self._unit.setObjectName("MetricUnit")

        layout.addWidget(self._lbl)
        row = QHBoxLayout()
        row.setSpacing(4)
        row.addWidget(self._val)
        row.addWidget(self._unit)
        row.addStretch()
        layout.addLayout(row)

        self.setStyleSheet(f"""
            QFrame#MetricCard {{
                background-color: {C['card']};
                border: 1px solid {C['border']};
                border-left: 3px solid {color};
                border-radius: 8px;
            }}
            QLabel#MetricLabel {{
                color: {C['muted']};
                font-size: 10px;
                font-weight: 600;
                letter-spacing: 1px;
            }}
            QLabel#MetricValue {{
                color: {C['text']};
                font-size: 22px;
                font-weight: 700;
                font-family: 'SF Mono', 'Consolas', monospace;
            }}
            QLabel#MetricUnit {{
                color: {color};
                font-size: 11px;
                font-weight: 600;
                padding-bottom: 4px;
            }}
        """)

    def set_value(self, text):
        self._val.setText(text)


# ─────────────────────────────────────────────────────────────────────────────
# Main Window
# ─────────────────────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Solar Tracker Monitor  v5.0")
        self.resize(1440, 900)

        self.data    = DataModel()
        self.signals = DataSignals()

        self.serial_port   = None
        self.reading       = False
        self.demo_running  = False
        self.start_time    = None

        self.measuring         = False
        self.measure_start_ts  = None
        self.measure_duration_s = 60

        self._build_ui()
        self._connect_signals()
        self._refresh_ports()

        self.plot_timer = QTimer(self)
        self.plot_timer.timeout.connect(self._update_plots)
        self.plot_timer.start(400)

        self.measure_timer = QTimer(self)
        self.measure_timer.timeout.connect(self._tick_measure)

    # ── UI ────────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.setStyleSheet(self._stylesheet())
        root = QWidget()
        self.setCentralWidget(root)
        root_lay = QVBoxLayout(root)
        root_lay.setContentsMargins(0, 0, 0, 0)
        root_lay.setSpacing(0)

        root_lay.addWidget(self._build_topbar())
        root_lay.addWidget(self._build_toolbar())

        body = QHBoxLayout()
        body.setContentsMargins(16, 12, 16, 12)
        body.setSpacing(14)
        body.addWidget(self._build_left_panel(), 0)
        body.addWidget(self._build_right_panel(), 1)
        root_lay.addLayout(body)

    def _build_topbar(self):
        bar = QFrame()
        bar.setObjectName("TopBar")
        bar.setFixedHeight(52)
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(20, 0, 20, 0)

        icon = QLabel("☀")
        icon.setStyleSheet(f"color: {C['yellow']}; font-size: 22px;")
        lay.addWidget(icon)

        title = QLabel("Solar Tracker Monitor")
        title.setObjectName("AppTitle")
        lay.addWidget(title)

        lay.addStretch()

        self.status_pill = QLabel("● NOT CONNECTED")
        self.status_pill.setObjectName("StatusPill")
        self.status_pill.setProperty("state", "disconnected")
        lay.addWidget(self.status_pill)
        return bar

    def _build_toolbar(self):
        bar = QFrame()
        bar.setObjectName("ToolBar")
        bar.setFixedHeight(48)
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(16, 0, 16, 0)
        lay.setSpacing(8)

        # Port
        self._tb_label(lay, "PORT")
        self.port_combo = QComboBox()
        self.port_combo.setFixedWidth(150)
        self.port_combo.setObjectName("TbCombo")
        lay.addWidget(self.port_combo)

        # Baud
        self._tb_label(lay, "BAUD")
        self.baud_combo = QComboBox()
        self.baud_combo.addItems(["9600", "115200"])
        self.baud_combo.setFixedWidth(90)
        self.baud_combo.setObjectName("TbCombo")
        lay.addWidget(self.baud_combo)

        self.connect_btn = self._tb_btn("Connect", C["accent"])
        self.connect_btn.clicked.connect(self._toggle_connect)
        lay.addWidget(self.connect_btn)

        self.demo_btn = self._tb_btn("▶  Demo", C["success"])
        self.demo_btn.clicked.connect(self._toggle_demo)
        lay.addWidget(self.demo_btn)

        rescan = self._tb_btn("↻  Rescan", C["border"])
        rescan.clicked.connect(self._refresh_ports)
        lay.addWidget(rescan)

        lay.addStretch()

        exp_all = self._tb_btn("⬇  Export Excel (all)", "#2D6A2D")
        exp_all.clicked.connect(lambda: self._export_xlsx(measurement_only=False))
        lay.addWidget(exp_all)

        exp_meas = self._tb_btn("⬇  Export Excel (measurement)", "#2D6A2D")
        exp_meas.clicked.connect(lambda: self._export_xlsx(measurement_only=True))
        lay.addWidget(exp_meas)

        exp_png = self._tb_btn("🖼  PNG", "#444")
        exp_png.clicked.connect(self._export_png)
        lay.addWidget(exp_png)

        return bar

    def _tb_label(self, lay, text):
        lbl = QLabel(text)
        lbl.setObjectName("TbLabel")
        lay.addWidget(lbl)

    def _tb_btn(self, text, bg):
        btn = QPushButton(text)
        btn.setObjectName("TbBtn")
        btn.setStyleSheet(f"""
            QPushButton#TbBtn {{
                background-color: {bg};
                color: {C['text']};
                border: none;
                border-radius: 6px;
                padding: 5px 12px;
                font-size: 12px;
                font-weight: 600;
            }}
            QPushButton#TbBtn:hover {{ background-color: #ffffff22; }}
            QPushButton#TbBtn:pressed {{ background-color: #ffffff11; }}
        """)
        return btn

    def _build_left_panel(self):
        panel = QFrame()
        panel.setObjectName("Panel")
        panel.setFixedWidth(240)
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        # INA219 metrics
        lay.addWidget(self._section_header("⚡  POWER METRICS"))
        self.mc_v  = MetricCard("Voltage",  "V",    C["blue"])
        self.mc_i  = MetricCard("Current",  "mA",   C["green"])
        self.mc_p  = MetricCard("Power",    "mW",   C["orange"])
        self.mc_e  = MetricCard("Energy",   "mWh",  C["purple"])
        for mc in (self.mc_v, self.mc_i, self.mc_p, self.mc_e):
            lay.addWidget(mc)

        # LDR sensors
        lay.addWidget(self._section_header("🔆  LDR SENSORS"))
        self.ldr_cards = {}
        for name, color in [("Top", C["red"]), ("Right", C["orange"]),
                             ("Bottom", C["cyan"]), ("Left", C["yellow"])]:
            mc = MetricCard(name, "", color)
            self.ldr_cards[name] = mc
            lay.addWidget(mc)

        # Servo
        lay.addWidget(self._section_header("🔄  SERVO POSITION"))
        self.mc_rot  = MetricCard("Rotation", "°", C["cyan"])
        self.mc_tilt = MetricCard("Tilt",     "°", C["yellow"])
        lay.addWidget(self.mc_rot)
        lay.addWidget(self.mc_tilt)

        # System
        lay.addWidget(self._section_header("🖥  SYSTEM"))
        info_card = QFrame()
        info_card.setObjectName("InfoCard")
        info_lay = QVBoxLayout(info_card)
        info_lay.setContentsMargins(12, 10, 12, 10)
        info_lay.setSpacing(4)
        self.serial_state  = QLabel("Serial: —")
        self.samples_label = QLabel("Samples: 0")
        self.uptime_label  = QLabel("Uptime: —")
        for lbl in (self.serial_state, self.samples_label, self.uptime_label):
            lbl.setObjectName("InfoText")
            info_lay.addWidget(lbl)
        clear_btn = QPushButton("Clear All Data")
        clear_btn.setObjectName("ClearBtn")
        clear_btn.clicked.connect(self._clear_data)
        info_lay.addWidget(clear_btn)
        lay.addWidget(info_card)

        lay.addStretch()
        return panel

    def _build_right_panel(self):
        panel = QWidget()
        lay = QVBoxLayout(panel)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        lay.addWidget(self._build_measure_bar())
        lay.addWidget(self._build_chart_panel(), 1)
        return panel

    def _build_measure_bar(self):
        bar = QFrame()
        bar.setObjectName("MeasureBar")
        lay = QHBoxLayout(bar)
        lay.setContentsMargins(16, 10, 16, 10)
        lay.setSpacing(12)

        lay.addWidget(QLabel("Duration (s):"))
        self.duration_spin = QSpinBox()
        self.duration_spin.setRange(1, 36000)
        self.duration_spin.setValue(60)
        self.duration_spin.setObjectName("SpinBox")
        lay.addWidget(self.duration_spin)

        for label, val in [("1 min", 60), ("5 min", 300), ("20 min", 1200)]:
            btn = QPushButton(label)
            btn.setObjectName("QuickBtn")
            btn.clicked.connect(lambda _, v=val: self.duration_spin.setValue(v))
            lay.addWidget(btn)

        self.measure_btn = QPushButton("▶  Start Measurement")
        self.measure_btn.setObjectName("MeasureBtn")
        self.measure_btn.clicked.connect(self._toggle_measure)
        lay.addWidget(self.measure_btn)

        self.progress = QProgressBar()
        self.progress.setObjectName("MeasureProgress")
        self.progress.setFixedWidth(160)
        lay.addWidget(self.progress)

        self.timer_label = QLabel("--:--  /  --:--")
        self.timer_label.setObjectName("TimerLabel")
        lay.addWidget(self.timer_label)

        lay.addStretch()

        self.stats_label = QLabel("Avg: —   Max: —   Min: —   Energy: —")
        self.stats_label.setObjectName("StatsLabel")
        lay.addWidget(self.stats_label)

        return bar

    def _build_chart_panel(self):
        card = QFrame()
        card.setObjectName("ChartCard")
        lay = QVBoxLayout(card)
        lay.setContentsMargins(12, 8, 12, 8)
        lay.setSpacing(0)

        hdr = QHBoxLayout()
        hdr.addWidget(self._section_header("📈  LIVE CHARTS"))
        hdr.addStretch()
        tab_row = QHBoxLayout()
        self._chart_tabs = {}
        for name in ("Power", "Voltage", "Current", "Energy", "LDR", "Position"):
            btn = QPushButton(name)
            btn.setObjectName("ChartTab")
            btn.setCheckable(True)
            btn.clicked.connect(lambda _, n=name: self._switch_tab(n))
            tab_row.addWidget(btn)
            self._chart_tabs[name] = btn
        tab_row.addStretch()
        hdr.addLayout(tab_row)
        lay.addLayout(hdr)

        self.fig = Figure(facecolor=C["card"], dpi=96)
        self.fig.subplots_adjust(hspace=0.55, top=0.93, bottom=0.08,
                                 left=0.08, right=0.97)
        self._setup_axes()
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setStyleSheet(f"background-color: {C['card']};")
        lay.addWidget(self.canvas, 1)

        # Activate Power tab by default
        self._switch_tab("Power")
        return card

    def _setup_axes(self):
        self.fig.clear()
        ax_cfg = [
            ("ax_p",  "Power",    C["orange"],  "mW"),
            ("ax_v",  "Voltage",  C["blue"],    "V"),
            ("ax_i",  "Current",  C["green"],   "mA"),
            ("ax_e",  "Energy",   C["purple"],  "mWh"),
        ]
        for idx, (name, title, color, unit) in enumerate(ax_cfg, 1):
            ax = self.fig.add_subplot(2, 2, idx)
            ax.set_facecolor(C["surface"])
            ax.tick_params(colors=C["muted"], labelsize=8)
            ax.spines[:].set_color(C["border"])
            ax.set_title(f"{title}  [{unit}]", color=C["text"],
                         fontsize=9, loc="left", pad=4)
            ax.grid(True, linestyle="--", alpha=0.2, color=C["border"])
            line, = ax.plot([], [], color=color, linewidth=1.5)
            ax.fill_between([], [], alpha=0.08, color=color)
            setattr(self, name, ax)
            setattr(self, f"line_{name[3:]}", line)

        # LDR axes (hidden by default)
        self.ax_ldr = self.fig.add_subplot(1, 1, 1)
        self.ax_ldr.set_visible(False)
        self.ax_ldr.set_facecolor(C["surface"])
        self.ax_ldr.tick_params(colors=C["muted"], labelsize=8)
        self.ax_ldr.spines[:].set_color(C["border"])
        self.ax_ldr.set_title("LDR Sensors", color=C["text"], fontsize=10, loc="left")
        self.ax_ldr.grid(True, linestyle="--", alpha=0.2, color=C["border"])
        self.ldr_lines = {}
        for name, color in [("Top", C["red"]), ("Right", C["orange"]),
                             ("Bottom", C["cyan"]), ("Left", C["yellow"])]:
            line, = self.ax_ldr.plot([], [], color=color, linewidth=1.5, label=name)
            self.ldr_lines[name] = line
        self.ax_ldr.legend(facecolor=C["surface"], edgecolor=C["border"],
                           labelcolor=C["text"], fontsize=8)

        # Position axes (hidden by default)
        self.ax_pos = self.fig.add_subplot(1, 1, 1)
        self.ax_pos.set_visible(False)
        self.ax_pos.set_facecolor(C["surface"])
        self.ax_pos.tick_params(colors=C["muted"], labelsize=8)
        self.ax_pos.spines[:].set_color(C["border"])
        self.ax_pos.set_title("Tracker Position  [°]", color=C["text"],
                              fontsize=10, loc="left")
        self.ax_pos.grid(True, linestyle="--", alpha=0.2, color=C["border"])
        self.pos_lines = {}
        for name, color in [("Rotation", C["cyan"]), ("Tilt", C["yellow"])]:
            line, = self.ax_pos.plot([], [], color=color, linewidth=1.5, label=name)
            self.pos_lines[name] = line
        self.ax_pos.legend(facecolor=C["surface"], edgecolor=C["border"],
                           labelcolor=C["text"], fontsize=8)

        self._current_tab = "Power"

    def _switch_tab(self, name):
        self._current_tab = name
        for n, btn in self._chart_tabs.items():
            btn.setChecked(n == name)

        multi_ax = [self.ax_p, self.ax_v, self.ax_i, self.ax_e]
        if name in ("LDR", "Position"):
            for ax in multi_ax:
                ax.set_visible(False)
            self.ax_ldr.set_visible(name == "LDR")
            self.ax_pos.set_visible(name == "Position")
            # reposition as full-frame subplot
            for ax in (self.ax_ldr, self.ax_pos):
                ax.set_position([0.08, 0.10, 0.89, 0.80])
        else:
            for ax in multi_ax:
                ax.set_visible(True)
            self.ax_ldr.set_visible(False)
            self.ax_pos.set_visible(False)
            # dim the 3 others, highlight the selected one
            highlight = {"Power": self.ax_p, "Voltage": self.ax_v,
                         "Current": self.ax_i, "Energy": self.ax_e}.get(name)
            for ax in multi_ax:
                for spine in ax.spines.values():
                    spine.set_color(C["accent"] if ax is highlight else C["border"])
                    spine.set_linewidth(2 if ax is highlight else 1)

        self.canvas.draw_idle()

    def _section_header(self, text):
        lbl = QLabel(text)
        lbl.setObjectName("SectionHeader")
        return lbl

    # ── Styles ────────────────────────────────────────────────────────────
    def _stylesheet(self):
        return f"""
        QMainWindow, QWidget {{
            background-color: {C['bg']};
            color: {C['text']};
            font-family: 'Segoe UI', 'SF Pro Display', 'Helvetica Neue', Arial, sans-serif;
        }}
        QFrame#TopBar {{
            background-color: {C['surface']};
            border-bottom: 1px solid {C['border']};
        }}
        QLabel#AppTitle {{
            color: {C['text']};
            font-size: 16px;
            font-weight: 700;
            margin-left: 6px;
        }}
        QLabel#StatusPill {{
            color: {C['muted']};
            font-size: 11px;
            font-weight: 600;
            padding: 3px 10px;
            border: 1px solid {C['border']};
            border-radius: 10px;
            background: {C['card']};
        }}
        QFrame#ToolBar {{
            background-color: {C['surface']};
            border-bottom: 1px solid {C['border']};
        }}
        QLabel#TbLabel {{
            color: {C['muted']};
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 1px;
        }}
        QComboBox#TbCombo {{
            background-color: {C['card']};
            color: {C['text']};
            border: 1px solid {C['border']};
            border-radius: 5px;
            padding: 3px 8px;
            font-size: 12px;
        }}
        QComboBox#TbCombo::drop-down {{ border: none; }}
        QFrame#Panel {{
            background-color: transparent;
        }}
        QLabel#SectionHeader {{
            color: {C['muted']};
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 1.5px;
            padding: 4px 2px 2px 2px;
        }}
        QFrame#InfoCard {{
            background-color: {C['card']};
            border: 1px solid {C['border']};
            border-radius: 8px;
        }}
        QLabel#InfoText {{
            color: {C['muted']};
            font-size: 11px;
        }}
        QPushButton#ClearBtn {{
            background-color: #2A1515;
            color: {C['red']};
            border: 1px solid {C['red']}44;
            border-radius: 5px;
            padding: 4px 8px;
            font-size: 11px;
            font-weight: 600;
        }}
        QPushButton#ClearBtn:hover {{ background-color: #3D1E1E; }}
        QFrame#MeasureBar {{
            background-color: {C['surface']};
            border: 1px solid {C['border']};
            border-radius: 10px;
        }}
        QLabel {{
            color: {C['text']};
            font-size: 12px;
        }}
        QSpinBox#SpinBox {{
            background-color: {C['card']};
            color: {C['text']};
            border: 1px solid {C['border']};
            border-radius: 5px;
            padding: 3px 6px;
        }}
        QPushButton#QuickBtn {{
            background-color: {C['card']};
            color: {C['muted']};
            border: 1px solid {C['border']};
            border-radius: 5px;
            padding: 4px 10px;
            font-size: 11px;
        }}
        QPushButton#QuickBtn:hover {{ color: {C['text']}; border-color: {C['blue']}; }}
        QPushButton#MeasureBtn {{
            background-color: {C['success']};
            color: white;
            border: none;
            border-radius: 6px;
            padding: 5px 14px;
            font-size: 12px;
            font-weight: 700;
        }}
        QPushButton#MeasureBtn:hover {{ background-color: #2EA043; }}
        QProgressBar#MeasureProgress {{
            border: 1px solid {C['border']};
            background-color: {C['card']};
            border-radius: 5px;
            height: 12px;
            text-align: center;
            font-size: 9px;
            color: {C['text']};
        }}
        QProgressBar#MeasureProgress::chunk {{
            background-color: {C['green']};
            border-radius: 4px;
        }}
        QLabel#TimerLabel {{
            color: {C['text']};
            font-family: 'SF Mono', 'Consolas', monospace;
            font-size: 13px;
            font-weight: 600;
        }}
        QLabel#StatsLabel {{
            color: {C['muted']};
            font-size: 11px;
        }}
        QFrame#ChartCard {{
            background-color: {C['card']};
            border: 1px solid {C['border']};
            border-radius: 10px;
        }}
        QPushButton#ChartTab {{
            background-color: transparent;
            color: {C['muted']};
            border: none;
            border-bottom: 2px solid transparent;
            padding: 4px 10px;
            font-size: 11px;
            font-weight: 600;
        }}
        QPushButton#ChartTab:checked {{
            color: {C['blue']};
            border-bottom: 2px solid {C['blue']};
        }}
        QPushButton#ChartTab:hover {{ color: {C['text']}; }}
        """

    def _connect_signals(self):
        self.signals.new_line.connect(self._handle_line)

    # ── Serial ────────────────────────────────────────────────────────────
    def _refresh_ports(self):
        if not SERIAL_AVAILABLE:
            self.port_combo.clear()
            self.port_combo.addItem("(pyserial not installed)")
            return
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.port_combo.clear()
        self.port_combo.addItems(ports if ports else ["(no ports)"])

    def _toggle_connect(self):
        if self.reading:
            self._disconnect()
        else:
            self._connect()

    def _connect(self):
        if not SERIAL_AVAILABLE:
            QMessageBox.critical(self, "Error", "pyserial not installed.")
            return
        port = self.port_combo.currentText()
        if not port or port.startswith("("):
            QMessageBox.warning(self, "No Port", "Select a valid serial port.")
            return
        baud = int(self.baud_combo.currentText())
        try:
            self.serial_port = serial.Serial(port, baud, timeout=1)
            time.sleep(1.5)
            self.reading   = True
            self.start_time = time.time()
            self.connect_btn.setText("Disconnect")
            self._set_status(f"● {port}  @  {baud} baud", "connected")
            self.serial_state.setText(f"Serial: {port}")
            threading.Thread(target=self._read_loop, daemon=True).start()
        except Exception as e:
            QMessageBox.critical(self, "Serial Error", str(e))

    def _disconnect(self):
        self.reading = False
        if self.serial_port:
            try: self.serial_port.close()
            except Exception: pass
        self.serial_port = None
        self.connect_btn.setText("Connect")
        self._set_status("● NOT CONNECTED", "disconnected")
        self.serial_state.setText("Serial: —")

    def _read_loop(self):
        while self.reading and self.serial_port:
            try:
                raw = self.serial_port.readline().decode("utf-8", errors="ignore").strip()
                if raw:
                    self.signals.new_line.emit(raw)
            except Exception:
                break
        self.reading = False

    def _toggle_demo(self):
        if self.demo_running:
            self.demo_running = False
            self.demo_btn.setText("▶  Demo")
            self._set_status("● DEMO STOPPED", "disconnected")
            self.serial_state.setText("Serial: Demo off")
        else:
            self.demo_running = True
            self.demo_btn.setText("■  Stop Demo")
            self.start_time = time.time()
            self._set_status("● DEMO RUNNING", "demo")
            self.serial_state.setText("Serial: Demo")
            threading.Thread(target=self._demo_loop, daemon=True).start()

    def _demo_loop(self):
        t = 0
        while self.demo_running:
            t += 1
            v = 5.0 + 0.5 * math.sin(t / 10) + random.uniform(-0.05, 0.05)
            i = 80 + 20 * math.sin(t / 8 + 1) + random.uniform(-2, 2)
            p = v * i
            top    = 500 + int(200 * math.sin(t / 15)) + random.randint(-10, 10)
            right  = 480 + int(200 * math.cos(t / 15)) + random.randint(-10, 10)
            bottom = 1023 - top  + random.randint(-10, 10)
            left   = 1023 - right + random.randint(-10, 10)
            rot    = int(90 + 30 * math.sin(t / 20))
            hoek   = int(90 + 20 * math.sin(t / 30))
            line = (f"T:{top} L:{left} R:{right} B:{bottom} "
                    f"U={v:.2f}V I={i:.1f}mA P={p:.1f}mW rot={rot} hoek={hoek}")
            self.signals.new_line.emit(line)
            time.sleep(1.0)

    def _set_status(self, text, state):
        self.status_pill.setText(text)
        colors = {
            "connected":    C["green"],
            "demo":         C["yellow"],
            "disconnected": C["muted"],
            "measuring":    C["orange"],
        }
        c = colors.get(state, C["muted"])
        self.status_pill.setStyleSheet(
            f"color: {c}; font-size: 11px; font-weight: 600; "
            f"padding: 3px 10px; border: 1px solid {c}44; "
            f"border-radius: 10px; background: {C['card']};"
        )

    # ── Data ──────────────────────────────────────────────────────────────
    def _handle_line(self, raw):
        m = PATTERN.search(raw)
        if not m:
            return
        top, left, right, bottom = (int(m.group(i)) for i in (1, 2, 3, 4))
        voltage    = float(m.group(5))
        current_mA = float(m.group(6))
        power_mW   = float(m.group(7))
        rot        = int(m.group(8))
        hoek       = int(m.group(9))

        if self.start_time is None:
            self.start_time = time.time()
        t_rel = time.time() - self.start_time

        self.data.add_sample(t_rel, top, left, right, bottom,
                             voltage, current_mA, power_mW, rot, hoek,
                             self.measuring)

        self.mc_v.set_value(f"{voltage:.2f}")
        self.mc_i.set_value(f"{current_mA:.1f}")
        self.mc_p.set_value(f"{power_mW:.1f}")
        self.mc_e.set_value(f"{self.data.total_energy_mWh:.4f}")

        self.ldr_cards["Top"].set_value(str(top))
        self.ldr_cards["Right"].set_value(str(right))
        self.ldr_cards["Bottom"].set_value(str(bottom))
        self.ldr_cards["Left"].set_value(str(left))

        self.mc_rot.set_value(str(rot))
        self.mc_tilt.set_value(str(hoek))

        n = len(self.data.all_rows)
        self.samples_label.setText(f"Samples: {n}")
        if self.data.session_start:
            up = int(time.time() - self.data.session_start)
            h, rem = divmod(up, 3600)
            m2, s = divmod(rem, 60)
            self.uptime_label.setText(
                f"Uptime: {h:02d}:{m2:02d}:{s:02d}" if h
                else f"Uptime: {m2:02d}:{s:02d}")

    def _update_plots(self):
        if not self.data.timestamps:
            return
        t = list(self.data.timestamps)

        def _refresh_ax(ax, line, data):
            line.set_data(t, data)
            if data:
                mn, mx = min(data), max(data)
                pad = (mx - mn) * 0.15 or 0.5
                ax.set_xlim(max(0, t[-1] - 120), t[-1] + 2)
                ax.set_ylim(mn - pad, mx + pad)
                # redraw fill
                for col in ax.collections:
                    col.remove()
                ax.fill_between(t, data, alpha=0.08,
                                color=line.get_color())

        _refresh_ax(self.ax_p, self.line_p, list(self.data.powers))
        _refresh_ax(self.ax_v, self.line_v, list(self.data.voltages))
        _refresh_ax(self.ax_i, self.line_i, list(self.data.currents))
        _refresh_ax(self.ax_e, self.line_e, list(self.data.energies))

        # LDR
        for name, dq in [("Top",    self.data.ldr_top),
                          ("Right",  self.data.ldr_right),
                          ("Bottom", self.data.ldr_bottom),
                          ("Left",   self.data.ldr_left)]:
            d = list(dq)
            self.ldr_lines[name].set_data(t, d)
        if self.data.ldr_top:
            vals = (list(self.data.ldr_top) + list(self.data.ldr_right) +
                    list(self.data.ldr_bottom) + list(self.data.ldr_left))
            mn, mx = min(vals), max(vals)
            pad = (mx - mn) * 0.1 or 10
            self.ax_ldr.set_xlim(max(0, t[-1] - 120), t[-1] + 2)
            self.ax_ldr.set_ylim(mn - pad, mx + pad)

        # Position
        for name, dq in [("Rotation", self.data.rot_pos),
                          ("Tilt",     self.data.hoek_pos)]:
            self.pos_lines[name].set_data(t, list(dq))
        if self.data.rot_pos:
            vals = list(self.data.rot_pos) + list(self.data.hoek_pos)
            mn, mx = min(vals), max(vals)
            pad = (mx - mn) * 0.1 or 5
            self.ax_pos.set_xlim(max(0, t[-1] - 120), t[-1] + 2)
            self.ax_pos.set_ylim(mn - pad, mx + pad)

        self.canvas.draw_idle()

    def _clear_data(self):
        self.data.clear()
        self.samples_label.setText("Samples: 0")
        self.mc_e.set_value("0.0000")

    # ── Measurement ───────────────────────────────────────────────────────
    def _toggle_measure(self):
        if self.measuring:
            self._stop_measure(manual=True)
        else:
            self._start_measure()

    def _start_measure(self):
        if not (self.reading or self.demo_running):
            QMessageBox.warning(self, "Not Connected",
                                "Connect to Arduino or start Demo mode first.")
            return
        self.measure_duration_s = int(self.duration_spin.value())
        self.measuring          = True
        self.measure_start_ts   = time.time()
        self.data.measure_rows  = []

        self.measure_btn.setText("■  Stop Measurement")
        self.measure_btn.setStyleSheet(
            f"background-color: {C['red']}; color: white; border: none; "
            f"border-radius: 6px; padding: 5px 14px; font-size: 12px; font-weight: 700;")
        self.progress.setMaximum(self.measure_duration_s)
        self.progress.setValue(0)
        self._set_status("● MEASURING", "measuring")
        self.measure_timer.start(200)

    def _tick_measure(self):
        if not self.measuring:
            return
        elapsed   = time.time() - self.measure_start_ts
        remaining = self.measure_duration_s - elapsed
        self.progress.setValue(min(int(elapsed), self.measure_duration_s))

        el  = time.strftime("%M:%S", time.gmtime(int(elapsed)))
        tot = time.strftime("%M:%S", time.gmtime(self.measure_duration_s))
        self.timer_label.setText(f"{el}  /  {tot}")

        if self.data.measure_rows:
            pw  = [r["power_mW"] for r in self.data.measure_rows]
            avg = sum(pw) / len(pw)
            e   = sum(r["power_mW"] * r["dt_s"] / 3600 for r in self.data.measure_rows)
            self.stats_label.setText(
                f"Avg: {avg:.1f} mW   Max: {max(pw):.1f} mW   "
                f"Min: {min(pw):.1f} mW   Energy: {e:.4f} mWh")

        if remaining <= 0:
            self._stop_measure(manual=False)

    def _stop_measure(self, manual):
        self.measuring = False
        self.measure_timer.stop()
        self.measure_btn.setText("▶  Start Measurement")
        self.measure_btn.setStyleSheet("")   # reset to stylesheet default
        reason = "stopped" if manual else "completed"

        n_s = len(self.data.measure_rows)
        self._set_status(
            f"● MEASUREMENT {reason.upper()}  ({n_s} samples)",
            "connected" if (self.reading or self.demo_running) else "disconnected")

        if self.data.measure_rows:
            pw  = [r["power_mW"] for r in self.data.measure_rows]
            e   = sum(r["power_mW"] * r["dt_s"] / 3600 for r in self.data.measure_rows)
            msg = (f"Measurement {reason}\n\n"
                   f"Samples   :  {n_s}\n"
                   f"Average   :  {sum(pw)/len(pw):.2f} mW\n"
                   f"Maximum   :  {max(pw):.2f} mW\n"
                   f"Minimum   :  {min(pw):.2f} mW\n"
                   f"Energy    :  {e:.5f} mWh\n\n"
                   "Export this measurement to Excel (with charts)?")
            if QMessageBox.question(self, "Measurement Done", msg,
                                    QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                self._export_xlsx(measurement_only=True)

    # ── Export ────────────────────────────────────────────────────────────
    def _export_xlsx(self, measurement_only):
        rows = self.data.measure_rows if measurement_only else self.data.all_rows
        if not rows:
            QMessageBox.information(self, "No Data",
                                    "No data available to export.")
            return
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = f"solar_measurement_{ts}.xlsx" if measurement_only \
               else f"solar_tracker_all_{ts}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", name, "Excel Files (*.xlsx)")
        if not path:
            return
        try:
            export_to_xlsx(rows, path)
            QMessageBox.information(self, "Exported",
                                    f"Saved with charts to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _export_png(self):
        if not self.data.timestamps:
            QMessageBox.information(self, "No Data", "No charts to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export PNG", "solar_tracker_chart.png", "PNG Files (*.png)")
        if path:
            self.fig.savefig(path, dpi=150, facecolor=C["card"])

    # ── Close ─────────────────────────────────────────────────────────────
    def closeEvent(self, event):
        self.reading      = False
        self.demo_running = False
        if self.serial_port:
            try: self.serial_port.close()
            except Exception: pass
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Solar Tracker Monitor")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()